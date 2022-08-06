'''Plots the conditions. All figures are shown simultaneously.
Part 1: Rizana river flow.
Part 2: see level in Koper.
Part 3: wind.
Part 4: air temperature at 2m height.
To save figures in Saved_figures directory, set savefigs=True'''

import openpyxl
from pathlib import Path
import datetime
import matplotlib.pyplot as plt
import numpy as np

savefigs = False # Whether it saves figures or not
mapa_za_shranjevanje_grafov = 'Saved_figures\ '[:-1]    # Where it saves the figures


# PART 1: RIZANA RIVER FLOW

my_sheet = '1.'
folder_with_xlsx_files = '..\Meritve_ARSO'
filenames = ['Kubed avg2008.xlsx', 'Kubed nov2008.xlsx', 'Kubed apr2009.xlsx']

datetime_min_max = {'Kubed apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=18, hour=23, minute=59), datetime.datetime(year=2009, month=4, day=21, hour=9, minute=5)),\
                    'Kubed avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=19, hour=23, minute=59), datetime.datetime(year=2008, month=8, day=22, hour=9, minute=5)),\
                    'Kubed nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=16, hour=23, minute=59), datetime.datetime(year=2008, month=11, day=19, hour=9, minute=5))\
                    }   # min and max date on plot

LM_min_max = {'Kubed apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=20, hour=7, minute=30), datetime.datetime(year=2009, month=4, day=21, hour=8, minute=14)),\
            'Kubed avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=21, hour=7, minute=9), datetime.datetime(year=2008, month=8, day=22, hour=7, minute=59)),\
            'Kubed nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=18, hour=6, minute=57), datetime.datetime(year=2008, month=11, day=19, hour=8, minute=22)) \
              } # min and max date of measurements with MSS90 probe

def rounder(t):
    if t.minute >= 30:
        return t.replace(second=0, microsecond=0, minute=0, hour=t.hour+1)
    else:
        return t.replace(second=0, microsecond=0, minute=0)

plt.figure(figsize=(10,7.1))
for ifn in range(len(filenames)):
    file_name = filenames[ifn]
    xslx_file = Path(folder_with_xlsx_files, file_name)

    wb_obj = openpyxl.load_workbook(xslx_file)

    this_datetime_min_max = datetime_min_max[file_name]
    this_datetime_min_max_LK = LM_min_max[file_name]
    used_datetime = []
    used_flow = []

    sheet = wb_obj[my_sheet]

    row = 15    # starting row (see xslx file, there is a header in rows 1-14)
    while True:
        try:
            this_datetime = sheet[f'C{row}'].value
            if this_datetime > this_datetime_min_max[0] and this_datetime < this_datetime_min_max[1]:
                used_datetime.append(this_datetime)
                used_flow.append(sheet[f'D{row}'].value)
        except: # EOF
            break
        row += 1



    plt.subplot(3, 1, ifn + 1)
    plt.plot([i for i in range(len(used_flow))], used_flow, color='tab:green')
    plt.xticks(ticks=[i for i in range(len(used_flow)) if rounder(used_datetime[i]).hour%4==0], labels=[f"{rounder(dt).hour:02d}" for dt in used_datetime if rounder(dt).hour%4 == 0], fontsize=14)
    plt.yticks(fontsize=14)
    plt.ylim(0,2)
    plt.xlim(0, len(used_flow) - 1)
    plt.title(f'{rounder(used_datetime[0]).day}.{rounder(used_datetime[0]).month}.{rounder(used_datetime[0]).year} {rounder(used_datetime[0]).hour:02d}h - {rounder(used_datetime[-1]).day}.{rounder(used_datetime[-1]).month}.{rounder(used_datetime[-1]).year} {rounder(used_datetime[-1]).hour:02d}h', fontsize=16)
    plt.ylabel(r'$[\mathrm{m}^3/\mathrm{s}]$', fontsize=14)
    plt.xlabel('Ura meritve', fontsize=16)  # 'Measurement time'
    starting_LK_measurements_fraction = (this_datetime_min_max_LK[0] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    ending_LK_measurements_fraction = (this_datetime_min_max_LK[1] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    plt.fill_betweenx(y=[0,2], x1=[starting_LK_measurements_fraction*(len(used_flow)-1), starting_LK_measurements_fraction*(len(used_flow)-1)], x2=[ending_LK_measurements_fraction*(len(used_flow)-1), ending_LK_measurements_fraction*(len(used_flow)-1)], color='C1', alpha=0.3)

plt.suptitle('Pretok Rižane v Kubedu', fontsize=18) # 'Rizana river flow in Kubed'
plt.tight_layout()
plt.subplots_adjust(top=0.85)
if savefigs:
    plt.savefig(mapa_za_shranjevanje_grafov + 'pretok_rizane.jpg', dpi=300)


# PART 2: SEE LEVEL IN KOPER

filename = 'Koper_vodostaj morja_MV.xlsx'
sheets = ['20.-23.8.2008', '17.-20.11.2008', '19.-22.4.2009']
datetime_min_max = {'19.-22.4.2009':(datetime.datetime(year=2009, month=4, day=18, hour=23, minute=59), datetime.datetime(year=2009, month=4, day=21, hour=9, minute=5)),\
                    '20.-23.8.2008':(datetime.datetime(year=2008, month=8, day=19, hour=23, minute=59), datetime.datetime(year=2008, month=8, day=22, hour=9, minute=5)),\
                    '17.-20.11.2008':(datetime.datetime(year=2008, month=11, day=16, hour=23, minute=59), datetime.datetime(year=2008, month=11, day=19, hour=9, minute=5))\
                    } # min and max date on plot

LM_min_max = {'19.-22.4.2009':(datetime.datetime(year=2009, month=4, day=20, hour=7, minute=30), datetime.datetime(year=2009, month=4, day=21, hour=8, minute=14)),\
            '20.-23.8.2008':(datetime.datetime(year=2008, month=8, day=21, hour=7, minute=9), datetime.datetime(year=2008, month=8, day=22, hour=7, minute=59)),\
            '17.-20.11.2008':(datetime.datetime(year=2008, month=11, day=18, hour=6, minute=57), datetime.datetime(year=2008, month=11, day=19, hour=8, minute=22)) \
              } # min and max date of measurements with MSS90 probe

def rounder(t):
    if t.minute >= 30 and t.hour < 23:
        return t.replace(second=0, microsecond=0, minute=0, hour=t.hour+1)
    elif t.minute >= 30:
        return t.replace(day=t.day+1, hour=0, second=0, microsecond=0, minute=0)
    else:
        return t.replace(second=0, microsecond=0, minute=0)


def color_title(labels, colors, textprops={'size': 'large'}, ax=None, y=1.013, precision=10 ** -2, xT=0, shift=0):
    """Creates a centered title with multiple colors. Needs manual adjustments once the figure is saved.
    Source:https://github.com/alexanderthclark/Matplotlib-for-Storytellers/blob/main/Python/color_title.py on 15.4.2022"""
    if ax == None:
        ax = plt.gca()
    plt.gcf().canvas.draw()
    transform = ax.transAxes  # use axes coords
    # initial params
    # xT... where the text ends in x-axis coords
    # shift... where the text starts
    # for text objects
    text = dict()
    while (np.abs(shift - (1 - xT)) > precision) and (shift <= xT):
        x_pos = shift
        for label, col in zip(labels, colors):
            try:
                text[label].remove()
            except KeyError:
                pass
            text[label] = ax.text(x_pos, y, label, transform=transform, ha='left', color=col, **textprops)
            x_pos = text[label].get_window_extent().transformed(transform.inverted()).x1
        xT = x_pos  # where all text ends
        shift += precision / 2  # increase for next iteration
        if x_pos > 1:  # guardrail
            break


fig, axs = plt.subplots(nrows=3, ncols=1, figsize=(10,7.5))
for isheet in range(len(sheets)):
    file_name = filename
    this_sheet = sheets[isheet]
    xslx_file = Path(folder_with_xlsx_files, file_name)

    wb_obj = openpyxl.load_workbook(xslx_file)

    this_datetime_min_max = datetime_min_max[this_sheet]
    this_datetime_min_max_LK = LM_min_max[this_sheet]
    used_datetime = []
    used_height = []
    used_w = []

    sheet = wb_obj[this_sheet]

    row = 15
    while True:
        try:
            this_datetime = sheet[f'C{row}'].value
            if this_datetime > this_datetime_min_max[0] and this_datetime < this_datetime_min_max[1]:
                used_datetime.append(this_datetime)
                used_height.append(sheet[f'D{row}'].value)
        except: # EOF
            break
        row += 1
    for iuh in range(len(used_height) - 1):
        used_w.append(10**3 * (used_height[iuh+1] - used_height[iuh])/(used_datetime[iuh+1] - used_datetime[iuh]).total_seconds()) # 10**-3 cm/s

    ch = 'b'
    cw = 'r'
    axs[isheet].plot([i for i in range(len(used_height))], used_height, color=ch, linestyle='-', label='Vodostaj [cm]')
    axs[isheet].set_xticks(ticks=[i for i in range(len(used_height)) if rounder(used_datetime[i]).hour%4==0])
    axs[isheet].set_xticklabels(labels=[f"{rounder(dt).hour:02d}" for dt in used_datetime if rounder(dt).hour%4 == 0], fontsize=14)

    axs[isheet].tick_params(axis='y', labelcolor=ch, labelsize=14)
    axs[isheet].set_ylim(165, 280)
    axs[isheet].set_xlim(0, len(used_height) - 1)
    axs[isheet].set_title(f'{rounder(used_datetime[0]).day}.{rounder(used_datetime[0]).month}.{rounder(used_datetime[0]).year} {rounder(used_datetime[0]).hour:02d}h - {rounder(used_datetime[-1]).day}.{rounder(used_datetime[-1]).month}.{rounder(used_datetime[-1]).year} {rounder(used_datetime[-1]).hour:02d}h', fontsize=16)
    axs[isheet].set_ylabel(r'[cm]', color=ch, fontsize=14)
    axs[isheet].set_xlabel('Ura meritve', fontsize=16)
    starting_LK_measurements_fraction = (this_datetime_min_max_LK[0] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    ending_LK_measurements_fraction = (this_datetime_min_max_LK[1] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    axs[isheet].fill_betweenx(y=[165, 280], x1=[starting_LK_measurements_fraction*(len(used_height)-1), starting_LK_measurements_fraction*(len(used_height)-1)], x2=[ending_LK_measurements_fraction*(len(used_height)-1), ending_LK_measurements_fraction*(len(used_height)-1)], color='C1', alpha=0.3)
    ax = axs[isheet].twinx()
    ax.plot([i + 1/2 for i in range(len(used_height) - 1)], used_w, color=cw, linestyle='-.')
    ax.set_ylabel(r'[$10^{-3}$ cm/s]', color=cw, fontsize=14)
    ax.set_ylim(-6, 6)
    ax.axhline(0, linestyle=':', linewidth=0.6, color=cw)
    ax.tick_params(axis='y', labelcolor=cw, labelsize=14)
    if isheet == 0:
        color_title(labels=['Višina vodostaja', ' in ', 'vertikalna hitrost gladine', ' na postaji Koper - Kapitanija'], colors=[ch, 'k', cw, 'k'], y=1.7, textprops={'size':18}, xT=0, shift=0)
plt.tight_layout()
plt.subplots_adjust(wspace=0)
# Legend was added manually
if savefigs:
    plt.savefig(mapa_za_shranjevanje_grafov + 'visina_vodostaja.jpg', dpi=300)



# PART 3: WIND
filenames = ['veter avg2008.xlsx', 'veter nov2008.xlsx', 'veter apr2009.xlsx']
sheets = ['Koper-Luka', 'Koper-Kapitanija']

datetime_min_max = {'veter apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=18, hour=23, minute=59), datetime.datetime(year=2009, month=4, day=21, hour=9, minute=5)),\
                    'veter avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=19, hour=23, minute=59), datetime.datetime(year=2008, month=8, day=22, hour=9, minute=5)),\
                    'veter nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=16, hour=23, minute=59), datetime.datetime(year=2008, month=11, day=19, hour=9, minute=5))\
                    } # min and max date on plot

LM_min_max = {'veter apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=20, hour=7, minute=30), datetime.datetime(year=2009, month=4, day=21, hour=8, minute=14)),\
            'veter avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=21, hour=7, minute=9), datetime.datetime(year=2008, month=8, day=22, hour=7, minute=59)),\
            'veter nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=18, hour=6, minute=57), datetime.datetime(year=2008, month=11, day=19, hour=8, minute=22)) \
              } # min and max date of measurements with MSS90 probe
colors = ['tab:blue', 'tab:red']

fig, axs = plt.subplots(nrows=3, ncols=1, figsize=(10,7.5))
linestyles = ['-', '-.']
labels = ['Luka', 'Kapitanija']
for ifn in range(len(filenames)):
    file_name = filenames[ifn]
    xslx_file = Path(folder_with_xlsx_files, file_name)

    wb_obj = openpyxl.load_workbook(xslx_file)

    this_datetime_min_max = datetime_min_max[file_name]
    this_datetime_min_max_LK = LM_min_max[file_name]
    for isheet in range(len(sheets)):
        used_datetime = []
        used_mean_wind = []

        sheet = wb_obj[sheets[isheet]]
        color = colors[isheet]

        row = 2
        while True:
            try:
                this_datetime = sheet[f'C{row}'].value
                if this_datetime > this_datetime_min_max[0] and this_datetime < this_datetime_min_max[1]:
                    used_datetime.append(this_datetime)
                    used_mean_wind.append(sheet[f'D{row}'].value)
            except: # EOF
                break
            row += 1



        axs[ifn].plot([i for i in range(len(used_mean_wind))], used_mean_wind, color=color, linestyle=linestyles[isheet], label=labels[isheet])
    axs[ifn].set_xticks(ticks=[i for i in range(len(used_mean_wind)) if (used_datetime[i]).hour%4==0 and (used_datetime[i]).minute==0])
    axs[ifn].set_xticklabels(labels=[f"{dt.hour:02d}" for dt in used_datetime if dt.hour%4 == 0 and dt.minute==0], fontsize=14)
    axs[ifn].set_ylim(0,7)
    axs[ifn].set_xlim(0, len(used_mean_wind) - 1)
    axs[ifn].set_title(f'{used_datetime[0].day}.{used_datetime[0].month}.{used_datetime[0].year} {used_datetime[0].hour:02d}h - {used_datetime[-1].day}.{used_datetime[-1].month}.{used_datetime[-1].year} {used_datetime[-1].hour:02d}h', fontsize=16)
    axs[ifn].set_ylabel(r'$[\mathrm{m}/\mathrm{s}]$', fontsize=14)
    axs[ifn].set_xlabel('Ura meritve', fontsize=16)
    axs[ifn].tick_params(axis='y', labelsize=14)
    starting_LK_measurements_fraction = (this_datetime_min_max_LK[0] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    ending_LK_measurements_fraction = (this_datetime_min_max_LK[1] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    axs[ifn].fill_betweenx(y=[0,7], x1=[starting_LK_measurements_fraction*(len(used_mean_wind)-1), starting_LK_measurements_fraction*(len(used_mean_wind)-1)], x2=[ending_LK_measurements_fraction*(len(used_mean_wind)-1), ending_LK_measurements_fraction*(len(used_mean_wind)-1)], color='C1', alpha=0.3)
    ax = axs[ifn].twinx()   # without this line color_title simply doesn't work!
    ax.set_yticks(ticks=[])
    if ifn == 0:
        axs[ifn].legend(fontsize=16, loc='upper left')
        color_title(labels=['Hitrost vetra na postajah', 'Koper - Luka', ' in ', 'Koper - Kapitanija'], colors=['k', colors[0], 'k', colors[1]], y=1.7, textprops={'size':18})#fontsize=14)

plt.tight_layout()
plt.subplots_adjust(wspace=0)
if savefigs:
    plt.savefig(mapa_za_shranjevanje_grafov + 'hitrost_vetra.jpg', dpi=300)



# PART 4: AIR TEMPERATURE


filenames = ['Koper_T2m_avg2008.xlsx', 'Koper_T2m_nov2008.xlsx', 'Koper_T2m_apr2009.xlsx']
sheets = ['Koper-Luka', 'Koper-Kapitanija']

datetime_min_max = {'Koper_T2m_apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=18, hour=23, minute=59), datetime.datetime(year=2009, month=4, day=21, hour=9, minute=5)),\
                    'Koper_T2m_avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=19, hour=23, minute=59), datetime.datetime(year=2008, month=8, day=22, hour=9, minute=5)),\
                    'Koper_T2m_nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=16, hour=23, minute=59), datetime.datetime(year=2008, month=11, day=19, hour=9, minute=5))\
                    }

LM_min_max = {'Koper_T2m_apr2009.xlsx':(datetime.datetime(year=2009, month=4, day=20, hour=7, minute=30), datetime.datetime(year=2009, month=4, day=21, hour=8, minute=14)),\
            'Koper_T2m_avg2008.xlsx':(datetime.datetime(year=2008, month=8, day=21, hour=7, minute=9), datetime.datetime(year=2008, month=8, day=22, hour=7, minute=59)),\
            'Koper_T2m_nov2008.xlsx':(datetime.datetime(year=2008, month=11, day=18, hour=6, minute=57), datetime.datetime(year=2008, month=11, day=19, hour=8, minute=22)) \
              }
colors = ['tab:blue', 'tab:red']
linestyles = ['-', '-.']
labels = ['Luka', 'Kapitanija']
fig, axs = plt.subplots(nrows=3, ncols=1, figsize=(10,7.5))
for ifn in range(len(filenames)):
    file_name = filenames[ifn]
    xslx_file = Path(folder_with_xlsx_files, file_name)

    wb_obj = openpyxl.load_workbook(xslx_file)

    this_datetime_min_max = datetime_min_max[file_name]
    this_datetime_min_max_LK = LM_min_max[file_name]
    for isheet in range(len(sheets)):
        used_datetime = []
        used_mean_T = []

        sheet = wb_obj[sheets[isheet]]
        color = colors[isheet]

        row = 2
        while True:
            try:
                this_datetime = sheet[f'C{row}'].value
                if this_datetime > this_datetime_min_max[0] and this_datetime < this_datetime_min_max[1]:
                    used_datetime.append(this_datetime)
                    used_mean_T.append(sheet[f'D{row}'].value)
            except: # EOF
                break
            row += 1



        axs[ifn].plot([i for i in range(len(used_mean_T))], used_mean_T, color=color, linestyle=linestyles[isheet], label=labels[isheet])
    axs[ifn].set_xticks(ticks=[i for i in range(len(used_mean_T)) if (used_datetime[i]).hour%4==0 and (used_datetime[i]).minute==0])
    axs[ifn].set_xticklabels(labels=[f"{dt.hour:02d}" for dt in used_datetime if dt.hour%4 == 0 and dt.minute==0], fontsize=14)
    axs[ifn].set_yticks([0, 5, 10, 15, 20, 25, 30, 35])
    axs[ifn].grid(linewidth=0.6, linestyle=':', color='gray')
    axs[ifn].set_xlim(0, len(used_mean_T) - 1)
    axs[ifn].set_title(f'{used_datetime[0].day}.{used_datetime[0].month}.{used_datetime[0].year} {used_datetime[0].hour:02d}h - {used_datetime[-1].day}.{used_datetime[-1].month}.{used_datetime[-1].year} {used_datetime[-1].hour:02d}h', fontsize=16)
    axs[ifn].set_ylabel(r'[°C]', fontsize=14)
    axs[ifn].set_xlabel('Ura meritve', fontsize=16)
    axs[ifn].tick_params(axis='y', labelsize=14)
    starting_LK_measurements_fraction = (this_datetime_min_max_LK[0] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    ending_LK_measurements_fraction = (this_datetime_min_max_LK[1] - used_datetime[0])/(used_datetime[-1] - used_datetime[0])
    axs[ifn].fill_betweenx(y=[-2, 35], x1=[starting_LK_measurements_fraction*(len(used_mean_T)-1), starting_LK_measurements_fraction*(len(used_mean_T)-1)], x2=[ending_LK_measurements_fraction*(len(used_mean_T)-1), ending_LK_measurements_fraction*(len(used_mean_T)-1)], color='C1', alpha=0.3)
    ax = axs[ifn].twinx()   # without this line color_title simply doesn't work!
    ax.set_yticks(ticks=[])
    if ifn == 0:
        axs[ifn].legend(fontsize=16, loc='upper left')
        color_title(labels=['Temperatura zraka na višini 2 m na postajah'], colors=['k'], y=1.7, textprops={'size':18})#fontsize=14)

axs[0].set_ylim([15,35])
axs[2].set_ylim([5,25])
axs[1].set_ylim([-2,18])
plt.tight_layout()
plt.subplots_adjust(wspace=0)
if savefigs:
    plt.savefig(mapa_za_shranjevanje_grafov + 'temperatura_zraka.jpg', dpi=300)




plt.show()